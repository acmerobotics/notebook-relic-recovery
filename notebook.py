#!python2

import os
import template
import openpyxl
import datetime


def append_line(s1, s2):
    return s1 + '\n' + s2

def findOccurences(s, ch):
    return [i for i, letter in enumerate(s) if letter == ch]

def substring_indexes(substring, string):

    last_found = -1  # Begin at -1 so the next position to search from is 0
    while True:
        # Find next index of substring, by starting after its last known position
        last_found = string.find(substring, last_found + 1)
        if last_found == -1:
            break  # All occurrences have been found
        yield last_found


def inIntervals(index, intervals):
    for interval in intervals:
        if  interval[0] < index < interval[1]:
            print "there is one in the intetrval"
            return True

    return False

def parse_response(response, path):


    begin = 0

    if 'Emma began to work on the Scouting App this week and was able to create a great deal of it' in response:
        print 'its the one'


    while '<' in response:

        codeStarts = list(substring_indexes('begin{lstlisting}', response))
        codeEnds = list(substring_indexes('end{lstlisting}', response))

        if not len(codeStarts) == len(codeEnds):
            raise Exception("you gonna have a problem" + response[:50])

        ignore = zip(codeStarts, codeEnds)

        start = response.find('<', begin) + 1
        if start == 0: break
        if inIntervals(start, ignore):
            begin = start+1
            continue
        end = response.find('>', start)
        begin = end + 1

        image = response[start:end]

        print image


        name = image.split('.')[0]

        figure = ''
        if ',' in image:
            file = image.split(',')[0].replace('_', '-')
            caption = image.split(',')[1]
            figure = template.figureTemplate.replace('path', path + file).replace('name', name).replace('captionr', caption)
        figure = figure + r'Figure \ref{fig:' + name + r'}'

        response = response[:start-1] + figure + response[end+1:]

        if response.find('<', end) == -1: break

    marks = findOccurences(response, '#')


    while '#' in response:
        start = findOccurences(response, '#')[0] + 1
        end = findOccurences(response, '#')[1]
        equation = response[start:end]
        name = ''
        result = ''
        if ',' in equation:
            name = equation.split(',')[0]
            result = template.eqTemplate.replace('value', equation.split(',')[1]).replace('name', name)
        else:
            name = equation


        result = result + r'Equation \ref{eq:' + name + '}'

        response = response[:start - 1] + result + response[end + 1:]

    return response

def getDate(week):
    return (datetime.date(2017, 8, 28) + datetime.timedelta(days=7*week)).strftime('%d %B %Y')

tex = open('./out/notebook.tex', 'w+')
tex.write(template.preamble)
tex.write(r'\begin{document}')
tex.write(r'\tableofcontents')
tex.write(r'\newpage')

dir = os.listdir('./in')
numWeeks = len(dir)

for i in range(numWeeks):
    goalsText = ''
    responseText = ''
    print 'week ', i + 1
    weekDir = dir[i]
    tex.write(template.weekBegin)
    tex.write(r'\subsection*{' + getDate(i) + r'}')
    weekList = os.listdir('./in/' + weekDir+ '/')
    if 'schedule.pdf' in weekList:
        tex.write(template.scheduleTemplate.replace('file', '../in/' + weekDir + '/' + 'schedule.pdf'))
    for file in weekList:
        if 'xlsx' not in file: continue
        first = file.lower()[0]
        team = 'Business'
        if first == 's': team = 'Software'
        elif first == 'h': team = 'Hardware'

        goalsText = append_line(goalsText, r'\subsection{' + team + r' Goals}')

        wb = openpyxl.load_workbook(filename='./in/' + weekDir + '/' + file);
        ws = wb.active

        goalNum = 1
        goal = ws['A' + str(goalNum + 3)].value
        while goal:
            goal = team[0] + str(goalNum) + ': ' + goal
            descip = ws['B' + str(goalNum + 3)].value
            response = ws['D' + str(goalNum + 3)].value

            try:
                goalsText = append_line(goalsText, r'\paragraph{' + goal + '} ' + descip)
            except Exception as e:
                print goal
                raise e

            response = parse_response(response, '../in/' + weekDir + '/')

            responseText = append_line(responseText, template.goalBeginTemplate.replace('goal', goal))
            responseText = append_line(responseText, response)

            goalNum += 1
            goal = ws['A' + str(goalNum + 3)].value

    tex.write(goalsText)
    tex.write(r'  \newpage  ' + '\n')
    tex.write(responseText)

tex.write(r'\end{document}')
tex.close()
